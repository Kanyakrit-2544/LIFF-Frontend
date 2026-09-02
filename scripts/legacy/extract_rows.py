#!/usr/bin/env python3
"""
ดึง "แถวดิบ" จากชีตขายจริง (Inner.xlsx) ออกมาเป็น JSON ให้ ETL ฝั่ง TS ใช้ต่อ

ต่างจาก profile_xlsx.py: อันนั้นดึงแค่ "สถิติ" (ไม่มี PII) เพื่อ commit ลง repo
อันนี้ดึง "ของจริง" (มี PII) — ผลลัพธ์ห้าม commit, ไหลผ่าน stdin ให้ import-legacy.ts เท่านั้น

    python3 scripts/legacy/extract_rows.py "raw input/Inner.xlsx" Inner2025,Inner2026 --limit 30

หัว/คอลัมน์ใช้พจนานุกรมชุดเดียวกับ profiler เป๊ะ (แก้ที่เดียวจบ)
ตรรกะธุรกิจ (dedupe/นับที่นั่ง/ยอด) อยู่ฝั่ง TS core (importReal.ts) ที่มีเทส — ที่นี่แค่ "อ่านไฟล์"
"""
import argparse
import json
import re
import sys
import datetime

import openpyxl

# หัวคอลัมน์ในชีต → field กลาง (เทียบแบบ "ขึ้นต้นด้วย" หลัง normalize) — ตรงกับ profile_xlsx.py
FIELD_HEADERS = {
    "no": ["no."],
    "paidAt": ["date", "วัน/เดือน/ปี ที่ชำระ"],
    "expiresAt": ["end date", "วันที่หมดเขต"],
    "slipNo": ["slip no.", "เลขที่สลิป"],
    "fullNameTh": ["ชื่อ - นามสกุล", "ชื่อ-นามสกุล"],
    "fullNameEn": ["name eng"],
    "nickname": ["ชื่อเล่น"],
    "age": ["อายุ"],
    "phone": ["เบอร์"],
    "social": ["fb/line/ig"],
    "email": ["email"],
    "amount": ["ยอดชำระ"],
    "saleRep": ["sale"],
    "note": ["หมายเหตุ", "note1", "crm"],
    "receipt": ["ใบเสร็จ", "ส่งใบเสร็จ", "ที่อยู่"],
}


def norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "")).strip()


def field_of(header):
    h = norm(header).lower()
    if not h:
        return None
    for field, prefixes in FIELD_HEADERS.items():
        for p in prefixes:
            if h.startswith(p):
                return field
    return None


def cell_json(v):
    """แปลงค่าเซลล์เป็นชนิดที่ JSON ถือได้ — วันที่เป็น ISO string ให้ฝั่ง TS แกะต่อ"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        # naive → ถือเป็น UTC midnight ให้ตรงกับ generator (Date.UTC) ไม่งั้นวันเพี้ยนตาม TZ เครื่อง
        return v.replace(tzinfo=datetime.timezone.utc).isoformat()
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day, tzinfo=datetime.timezone.utc).isoformat()
    if isinstance(v, (int, float, str, bool)):
        return v
    return norm(v)


def extract_sheet(ws, name, limit_state):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_i = next(
        (i for i, r in enumerate(rows[:8]) if any("นามสกุล" in norm(c) for c in r)),
        None,
    )
    if header_i is None:
        raise SystemExit(f"หา header row (ที่มี 'นามสกุล') ในชีต {name} ไม่เจอ")
    header = rows[header_i]

    field_cols, course_cols = {}, []
    for j, cell in enumerate(header):
        h = norm(cell)
        if not h:
            continue
        f = field_of(h)
        if f:
            field_cols.setdefault(f, j)
        else:
            course_cols.append((j, h))

    out_rows = []
    for offset, r in enumerate(rows[header_i + 1:]):
        if limit_state["remaining"] is not None and limit_state["remaining"] <= 0:
            break
        # เลขแถวจริงตามที่ Excel แสดง (1-based): header_i(0-based)+offset+2
        row_number = header_i + offset + 2
        fields = {}
        for f, j in field_cols.items():
            val = cell_json(r[j]) if j < len(r) else None
            if val is not None and val != "":
                fields[f] = val
        courses = []
        for j, label in course_cols:
            val = cell_json(r[j]) if j < len(r) else None
            if val is not None and val != "":
                courses.append({"label": label, "value": val})
        out_rows.append({"rowNumber": row_number, "fields": fields, "courses": courses})
        if limit_state["remaining"] is not None:
            limit_state["remaining"] -= 1

    return {"sheet": name, "rows": out_rows}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("sheets", help="ชื่อชีตคั่นด้วย comma เช่น Inner2025,Inner2026")
    ap.add_argument("--limit", type=int, default=None, help="จำนวนแถวรวมสูงสุด (ข้ามชีต)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    limit_state = {"remaining": args.limit}
    sheets = []
    for name in [s.strip() for s in args.sheets.split(",") if s.strip()]:
        if name not in wb.sheetnames:
            raise SystemExit(f"ไม่พบชีต {name} — มี: {', '.join(wb.sheetnames)}")
        sheets.append(extract_sheet(wb[name], name, limit_state))

    json.dump({"source": args.file, "sheets": sheets}, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
